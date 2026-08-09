from __future__ import annotations

import argparse
import csv
import json
import re
import xml.etree.ElementTree as ET
from collections import defaultdict
from copy import deepcopy
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATE_TAG = "20260731"
SLUG = "urban_follower_excel"

DEFAULT_EXCEL = ROOT / "Urban-Follower.xlsx"
DEFAULT_NETWORK = ROOT / "network/real_world_gaepo_modi/modi.inpx"
BASE_MAPPING = ROOT / "evaluation/real_world_modi_control/control_mapping.json"
BASE_DETECTOR = ROOT / "evaluation/real_world_modi_control/detector_local_mapping.json"
BASE_TUNING = "real_world_modi_pstack_vsl_rollout_vissimdsd_20260725.json"
CALIBRATION = ROOT / "evaluation/calibration/real_world_modi_control_v0_20260719.json"

OUT_DIR = ROOT / f"evaluation/real_world_modi_control_{SLUG}_{DATE_TAG}"
CONFIG_PATH = ROOT / f"evaluation/configs/real_world_modi_pstack_{SLUG}_{DATE_TAG}.json"
GENERATED_VBS_PATH = ROOT / f"evaluation/generated/real_world_modi_control_config_{SLUG}_{DATE_TAG}.vbs"
WRAPPER_PATH = ROOT / f"scripts/run_real_world_single_watchdog_{SLUG}.ps1"
REPORT_PATH = ROOT / f"outputs/real_world_{SLUG}_adapter_{DATE_TAG}.md"

DEFAULT_STORAGE_PER_MOVEMENT_VEH = 110.0
DEFAULT_CONTROLLED_UF_IDS = (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 13, 14, 15, 16)


@dataclass(frozen=True)
class ExcelHead:
    uf_id: int
    sc_no: int
    sg_no: int
    movement: str
    signal_head_no: int
    link_no: int
    lane_no: int
    pos_m: float

    @property
    def signal_id(self) -> str:
        return f"UF{self.uf_id}"

    @property
    def movement_key(self) -> str:
        return f"{self.signal_id}_{self.movement}_to_out"

    @property
    def phase(self) -> str:
        return f"{self.signal_id}_{phase_id(self.movement)}"

    @property
    def storage_link(self) -> str:
        return f"{self.signal_id}_{self.movement}_out"


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT)).replace("/", "\\")
    except ValueError:
        return str(path)


def read_json(path: Path) -> dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError(f"{path} must contain a JSON object")
    return data


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")


def first_int(text: str | None) -> int | None:
    match = re.search(r"-?\d+", str(text or ""))
    return int(match.group(0)) if match else None


def parse_lane_ref(text: str | None) -> tuple[int, int]:
    nums = [int(value) for value in re.findall(r"-?\d+", str(text or ""))]
    if not nums:
        raise ValueError(f"Cannot parse signal-head lane reference: {text!r}")
    link = nums[0]
    lane = nums[1] if len(nums) > 1 else 1
    return link, lane


def parse_sg_ref(text: str | None) -> tuple[int, int] | None:
    nums = [int(value) for value in re.findall(r"-?\d+", str(text or ""))]
    if len(nums) < 2:
        return None
    return nums[0], nums[1]


def parse_movement(text: Any) -> tuple[int, str]:
    raw = str(text or "").strip()
    nums = [int(value) for value in re.findall(r"-?\d+", raw)]
    if not nums:
        raise ValueError(f"Cannot parse movement label from {raw!r}")
    tail = raw.split(":", 1)[1].strip() if ":" in raw else raw
    match = re.search(r"[A-Za-z]+", tail)
    movement = match.group(0).upper() if match else f"SG{nums[0]}"
    return nums[0], movement


def phase_id(movement: str) -> str:
    movement = movement.upper()
    if movement.startswith(("E", "W")):
        return "p2"
    if movement.startswith(("N", "S")):
        return "p1"
    return "p1"


def axis_id(movement: str) -> str:
    return "EW" if phase_id(movement) == "p2" else "NS"


def parse_network(path: Path) -> tuple[dict[int, dict[str, Any]], dict[int, dict[str, Any]]]:
    root = ET.parse(path).getroot()
    controllers: dict[int, dict[str, Any]] = {}
    for sc in root.iter("signalController"):
        no = first_int(sc.get("no"))
        if no is None:
            continue
        controllers[no] = {
            "name": sc.get("name", ""),
            "signal_groups": {
                int(first_int(sg.get("no"))): sg.get("name", "")
                for sg in sc.findall("./sgs/signalGroup")
                if first_int(sg.get("no")) is not None
            },
        }

    heads: dict[int, dict[str, Any]] = {}
    for head in root.iter("signalHead"):
        no = first_int(head.get("no"))
        if no is None:
            continue
        sg_ref = parse_sg_ref(head.get("sg"))
        link_no, lane_no = parse_lane_ref(head.get("lane"))
        try:
            pos_m = float(str(head.get("pos", "0")).strip())
        except ValueError:
            pos_m = 0.0
        heads[no] = {
            "sc_no": sg_ref[0] if sg_ref else None,
            "sg_no": sg_ref[1] if sg_ref else None,
            "link_no": link_no,
            "lane_no": lane_no,
            "pos_m": pos_m,
            "name": head.get("name", ""),
        }
    return controllers, heads


def read_excel_heads(excel_path: Path, network_path: Path) -> tuple[list[ExcelHead], list[str], dict[int, dict[str, Any]]]:
    controllers, network_heads = parse_network(network_path)
    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    column = {name: idx for idx, name in enumerate(headers)}
    required = ["Urban Follower ID", "SC", "No.", "Signal Head"]
    missing = [name for name in required if name not in column]
    if missing:
        raise ValueError(f"{excel_path} is missing required columns: {missing}")

    rows: list[ExcelHead] = []
    warnings: list[str] = []
    for row_no, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value is not None for value in values):
            continue
        uf_raw = values[column["Urban Follower ID"]]
        sc_raw = values[column["SC"]]
        movement_raw = values[column["No."]]
        head_raw = values[column["Signal Head"]]
        if uf_raw is None or sc_raw is None or head_raw is None:
            warnings.append(f"row {row_no}: skipped incomplete row")
            continue
        uf_id = int(float(uf_raw))
        sc_no = int(float(sc_raw))
        signal_head_no = int(float(head_raw))
        sg_no, movement = parse_movement(movement_raw)

        if sc_no not in controllers:
            raise ValueError(f"row {row_no}: SC {sc_no} does not exist in {network_path}")
        if signal_head_no not in network_heads:
            raise ValueError(f"row {row_no}: signal head {signal_head_no} does not exist in {network_path}")
        head = network_heads[signal_head_no]
        if int(head["sc_no"]) != sc_no:
            raise ValueError(
                f"row {row_no}: signal head {signal_head_no} belongs to SC {head['sc_no']}, "
                f"not workbook SC {sc_no}"
            )
        if int(head["sg_no"]) != sg_no:
            raise ValueError(
                f"row {row_no}: signal head {signal_head_no} belongs to SG {head['sg_no']}, "
                f"not workbook SG {sg_no}"
            )
        sg_name = str(controllers[sc_no]["signal_groups"].get(sg_no, "")).upper()
        if sg_name and sg_name != movement:
            warnings.append(
                f"row {row_no}: workbook movement {movement} differs from modi.inpx SG name {sg_name} "
                f"for SC {sc_no} SG {sg_no}"
            )

        rows.append(
            ExcelHead(
                uf_id=uf_id,
                sc_no=sc_no,
                sg_no=sg_no,
                movement=movement,
                signal_head_no=signal_head_no,
                link_no=int(head["link_no"]),
                lane_no=int(head["lane_no"]),
                pos_m=float(head["pos_m"]),
            )
        )

    if not rows:
        raise ValueError(f"{excel_path} did not contain any usable urban follower rows")

    sc_by_uf: dict[int, set[int]] = defaultdict(set)
    for row in rows:
        sc_by_uf[row.uf_id].add(row.sc_no)
    multi_sc = {uf: sorted(scs) for uf, scs in sc_by_uf.items() if len(scs) > 1}
    if multi_sc:
        raise ValueError(f"Each Urban Follower ID must map to one SC; found {multi_sc}")

    return sorted(rows, key=lambda x: (x.uf_id, x.sg_no, x.signal_head_no)), warnings, controllers


def group_heads(rows: list[ExcelHead]) -> dict[int, list[ExcelHead]]:
    grouped: dict[int, list[ExcelHead]] = defaultdict(list)
    for row in rows:
        grouped[row.uf_id].append(row)
    return dict(sorted(grouped.items()))


def parse_controlled_uf_ids(text: str) -> set[int]:
    if not text.strip():
        return set(DEFAULT_CONTROLLED_UF_IDS)
    return {int(value) for value in re.findall(r"-?\d+", text)}


def filtered_rows(rows: list[ExcelHead], uf_ids: set[int]) -> list[ExcelHead]:
    return [row for row in rows if row.uf_id in uf_ids]


def unique_movements(rows: list[ExcelHead]) -> list[ExcelHead]:
    selected: dict[tuple[int, str], ExcelHead] = {}
    for row in rows:
        selected.setdefault((row.sg_no, row.movement), row)
    return [selected[key] for key in sorted(selected)]


def build_network_override(rows: list[ExcelHead]) -> dict[str, Any]:
    grouped = group_heads(rows)
    signals = [f"UF{uf_id}" for uf_id in grouped]
    grid_node_legs: dict[str, Any] = {}
    urban_movements: dict[str, Any] = {}
    boundary_in: list[str] = []
    boundary_out: list[str] = []
    storage: dict[str, float] = {}

    for uf_id, uf_rows in grouped.items():
        sid = f"UF{uf_id}"
        grid_node_legs[sid] = {}
        for movement_row in unique_movements(uf_rows):
            code = movement_row.movement
            in_link = f"in_{sid}_{code}"
            out_link = f"out_{sid}_{code}"
            storage_link = movement_row.storage_link
            grid_node_legs[sid][code] = {
                "type": "boundary",
                "in": in_link,
                "out": out_link,
                "out_link": storage_link,
            }
            boundary_in.append(in_link)
            boundary_out.append(out_link)
            storage[storage_link] = DEFAULT_STORAGE_PER_MOVEMENT_VEH
            urban_movements[movement_row.movement_key] = {
                "intersection": sid,
                "approach": code,
                "exit": "out",
                "beta": 1.0,
                "signal": sid,
                "phase": movement_row.phase,
                "origin": in_link,
                "destination": storage_link,
                "receiving_link": storage_link,
                "kind": "boundary_in",
            }

    return {
        "signals": signals,
        "uncontrolled_nodes": [],
        "urban_links": [],
        "grid_node_legs": grid_node_legs,
        "urban_movements": urban_movements,
        "boundary_in_links": boundary_in,
        "boundary_out_links": boundary_out,
        "urban_link_storage_veh": storage,
        "on_ramp_to_movement": {
            "R_D_W": [],
            "R_D_E": [],
            "R_F_W": [],
            "R_F_E": [],
        },
        "off_ramp_to_movement": {
            "OR_D_W": [],
            "OR_D_E": [],
            "OR_F_W": [],
            "OR_F_E": [],
        },
    }


def signal_group_summary(rows: list[ExcelHead]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    by_group: dict[tuple[int, str], list[ExcelHead]] = defaultdict(list)
    for row in rows:
        by_group[(row.sg_no, row.movement)].append(row)
    for (sg_no, movement), group in sorted(by_group.items()):
        out.append(
            {
                "sg_no": sg_no,
                "movement": movement,
                "axis": axis_id(movement),
                "phase": phase_id(movement),
                "head_count": len(group),
                "links": sorted({row.link_no for row in group}),
                "signal_heads": [row.signal_head_no for row in group],
            }
        )
    return out


def build_control_mapping(
    base: dict[str, Any],
    rows: list[ExcelHead],
    all_rows: list[ExcelHead],
    excel_path: Path,
    network_path: Path,
    detector_path: Path,
) -> dict[str, Any]:
    grouped = group_heads(rows)
    all_grouped = group_heads(all_rows)
    monitoring_grouped = {uf_id: group for uf_id, group in all_grouped.items() if uf_id not in grouped}
    out = deepcopy(base)
    out["created_at"] = f"2026-07-31"
    out["network"] = str(network_path)
    out["description"] = (
        "Urban follower adapter generated from Urban-Follower.xlsx and the marked UF1-UF19 image. "
        "Freeway VSL and ramp metering are copied from the validated real-world mapping; urban signal "
        "actuation is exposed for the selected UF subset while physical VISSIM SC numbers come from the workbook."
    )
    out["detector_mapping_json"] = str(detector_path)
    out["urban_signal_selector"] = SLUG
    out["source_excel"] = str(excel_path)
    out["source_network"] = str(network_path)
    out["controlled_urban_follower_ids"] = sorted(grouped)
    out["monitoring_only_urban_follower_ids"] = sorted(monitoring_grouped)
    out["monitoring_only_signal_controllers"] = [group[0].sc_no for group in monitoring_grouped.values()]
    out["signals"] = []
    for uf_id, uf_rows in grouped.items():
        sc_no = uf_rows[0].sc_no
        out["signals"].append(
            {
                "id": f"UF{uf_id}",
                "sc_no": sc_no,
                "urban_follower_id": uf_id,
                "source": "Urban-Follower.xlsx",
                "signal_group_filter": sorted({row.sg_no for row in uf_rows}),
                "signal_heads": [row.signal_head_no for row in uf_rows],
                "observed_links": sorted({row.link_no for row in uf_rows}),
                "signal_groups": signal_group_summary(uf_rows),
                "phase_map": {
                    "major_axis": "east_west",
                    "minor_axis": "north_south",
                    "major_phase": "p2",
                    "minor_phase": "p1",
                },
            }
        )
    return out


def add_unique(values: list[Any], value: Any) -> None:
    if value not in values:
        values.append(value)


def build_detector_mapping(base: dict[str, Any], rows: list[ExcelHead], controlled_uf_ids: set[int]) -> dict[str, Any]:
    grouped = group_heads(rows)
    out = deepcopy(base)
    out["mapping_version"] = f"real_world_modi_{SLUG}_v1_{DATE_TAG}"
    out["description"] = (
        "Local observation mapping generated from Urban-Follower.xlsx. Link counts at the selected "
        "signal-head stop-line links are split to UF movement queues by the number of selected heads "
        "per link/movement. Base freeway, off-ramp, and ramp-meter observations are preserved."
    )
    observable = {int(value) for value in out.get("observable_links", []) if str(value).strip().isdigit()}
    link_to_origins: dict[str, list[str]] = {
        str(key): [str(v) for v in value]
        for key, value in dict(out.get("link_to_origins", {})).items()
        if isinstance(value, list)
    }
    link_to_movements: dict[str, list[dict[str, Any]]] = {
        str(key): [dict(item) for item in value if isinstance(item, dict)]
        for key, value in dict(out.get("link_to_movements", {})).items()
        if isinstance(value, list)
    }

    by_link_movement: dict[tuple[int, str], list[ExcelHead]] = defaultdict(list)
    for row in rows:
        observable.add(row.link_no)
        by_link_movement[(row.link_no, row.movement_key)].append(row)

    for (link_no, movement_key), group in sorted(by_link_movement.items()):
        sample = group[0]
        link_key = str(link_no)
        origins = link_to_origins.setdefault(link_key, [])
        add_unique(origins, sample.storage_link)
        entries = link_to_movements.setdefault(link_key, [])
        existing = {
            str(item.get("movement", ""))
            for item in entries
            if isinstance(item, dict)
        }
        if movement_key not in existing:
            entries.append(
                {
                    "movement": movement_key,
                    "weight": float(len(group)),
                    "source": "Urban-Follower.xlsx",
                    "signal": sample.signal_id,
                    "phase": sample.phase,
                    "axis": axis_id(sample.movement),
                    "sc_no": sample.sc_no,
                    "urban_follower_id": sample.uf_id,
                    "signal_group_no": sample.sg_no,
                    "movement_label": sample.movement,
                    "signal_heads": [row.signal_head_no for row in group],
                }
            )

    agents: dict[str, Any] = {}
    for agent_id, spec in dict(base.get("agents", {})).items():
        if not isinstance(spec, dict):
            continue
        if str(spec.get("kind", "")).startswith("urban"):
            continue
        agents[str(agent_id)] = deepcopy(spec)

    for uf_id, uf_rows in grouped.items():
        sid = f"UF{uf_id}"
        control_enabled = uf_id in controlled_uf_ids
        agents[f"U_{sid}"] = {
            "kind": "urban" if control_enabled else "urban_monitor",
            "signal": sid,
            "sc_no": uf_rows[0].sc_no,
            "urban_follower_id": uf_id,
            "control_enabled": control_enabled,
            "monitoring_only": not control_enabled,
            "visible_links": [str(v) for v in sorted({row.link_no for row in uf_rows})],
            "visible_movements": sorted({row.movement_key for row in uf_rows}),
            "visible_ramps": [],
            "visible_off_ramps": [],
            "signal_group_filter": sorted({row.sg_no for row in uf_rows}),
            "signal_heads": [row.signal_head_no for row in uf_rows],
        }

    out["observable_links"] = sorted(observable)
    out["link_to_origins"] = link_to_origins
    out["link_to_movements"] = link_to_movements
    out["agents"] = agents
    out["guardrails"] = {
        **dict(base.get("guardrails", {})),
        "urban_signal_actuation": "urban_follower_excel_uf1_uf19",
        "distributed_urban_signal_players": {
            "enabled": True,
            "selector": SLUG,
            "controlled_urban_followers": [f"UF{uf_id}" for uf_id in grouped if uf_id in controlled_uf_ids],
            "monitoring_only_urban_followers": [f"UF{uf_id}" for uf_id in grouped if uf_id not in controlled_uf_ids],
            "controlled_signal_controllers": [
                group[0].sc_no for uf_id, group in grouped.items() if uf_id in controlled_uf_ids
            ],
            "note": "Only workbook-listed SC/SG pairs enter the COM signal-control filter.",
        },
    }
    return out


def build_tuning_config(
    network_override: dict[str, Any],
    mapping_path: Path,
    detector_path: Path,
    rows: list[ExcelHead],
    all_rows: list[ExcelHead],
    excel_path: Path,
    network_path: Path,
) -> dict[str, Any]:
    grouped = group_heads(rows)
    all_grouped = group_heads(all_rows)
    signals = network_override["signals"]
    monitoring_signals = [f"UF{uf_id}" for uf_id in all_grouped if uf_id not in grouped]
    return {
        "extends": BASE_TUNING,
        "name": f"real_world_modi_pstack_{SLUG}_{DATE_TAG}",
        "description": (
            "Distributed urban-follower adapter generated from the user's UF1-UF19 image and "
            "Urban-Follower.xlsx. Uses the user-modified modi.inpx as the VISSIM network."
        ),
        "mapping_json": str(mapping_path),
        "detector_mapping_json": str(detector_path),
        "source_network": str(network_path),
        "source_excel": str(excel_path),
        "urban_signal_selector": SLUG,
        "config_overrides": {
            "network": network_override,
            "mpc": {
                "follower_solver_mode": "distributed",
                "distributed_coupling_tol": 0.05,
                "stackelberg_allocation_mode": "simplified",
                "leader_candidate_count": 9,
                "leader_refinement_candidate_count": 5,
                "max_nash_iter": 4,
                "stackelberg_prefilter_top_k": 3,
                "stackelberg_prefilter_local_top_k": 3,
            },
            "urban_follower": {
                "allocation_pso_particles": 6,
                "allocation_pso_iterations": 8,
                "max_offset_step": 15.0,
                "offset_smoothness_weight": 0.08,
                "green_smoothness_weight": 0.08,
            },
        },
        "actuation": {
            "active_lever_mask": {
                "enabled": False,
                "allowed_signals": signals,
                "monitoring_only_signals": monitoring_signals,
            },
            "real_world_signal_control": {
                "enabled": True,
                "apply_to_no_control": False,
            },
            "real_world_signal_sg_filter": {
                str(group[0].sc_no): sorted({row.sg_no for row in group})
                for group in grouped.values()
            },
        },
        "safety": {
            "original_files_unchanged": True,
            "source_network_expected": str(network_path),
            "copy_only_outputs": [str(mapping_path), str(detector_path)],
        },
        "notes": [
            "The generated VBS config points the runner at workbook-listed physical SC numbers.",
            "The active signal-control follower list is generated from --controlled-uf-ids; all remaining workbook UFs stay monitoring-only.",
            "Signal groups are filtered by workbook-listed SG numbers before ContrByCOM and SigState are applied.",
            "The runner still controls signal groups, not individual signal heads; selected and unselected heads sharing the same SG cannot be separated at runtime.",
        ],
    }


def signal_sg_filter(rows: list[ExcelHead]) -> str:
    grouped_by_sc: dict[int, set[int]] = defaultdict(set)
    for row in rows:
        grouped_by_sc[row.sc_no].add(row.sg_no)
    return ";".join(
        f"{sc_no}:{'|'.join(str(sg) for sg in sorted(sgs))}"
        for sc_no, sgs in sorted(grouped_by_sc.items())
    )


def signal_head_filter(rows: list[ExcelHead]) -> str:
    grouped_by_sc: dict[int, list[int]] = defaultdict(list)
    for row in rows:
        grouped_by_sc[row.sc_no].append(row.signal_head_no)
    return ";".join(
        f"{sc_no}:{'|'.join(str(head) for head in heads)}"
        for sc_no, heads in sorted(grouped_by_sc.items())
    )


def write_generated_vbs(path: Path, mapping: dict[str, Any], detector: dict[str, Any], rows: list[ExcelHead], detector_path: Path) -> None:
    fw = mapping["freeway_model_links"]
    ramp_meters = list(mapping.get("ramp_meters", []))
    signals = list(mapping.get("signals", []))
    lines = [
        "' Generated by scripts/generate_urban_follower_excel_adapter.py",
        "RW_SCHEMA_VERSION = 1",
        'RW_FREEWAY_LINKS = "2,26"',
        'RW_FREEWAY_INPUT_LINKS = "26,74"',
        "RW_CLASSIFY_UNMATCHED_AS_URBAN = True",
    ]
    for model_link, var in (("FW_E", "RW_FW_E"), ("FW_W", "RW_FW_W")):
        spec = fw[model_link]
        lines.extend(
            [
                f"{var}_LINK = {int(spec['physical_link'])}",
                f"{var}_LENGTH_M = {float(spec['length_m']):.6f}",
                f"{var}_LANES = {int(spec['lanes'])}",
                f'{var}_SEG_BOUNDS = "' + ",".join(f"{float(v):.6f}" for v in spec["segment_bounds_m"]) + '"',
                f'{var}_SEG_LENGTHS_KM = "' + ",".join(f"{float(v):.6f}" for v in spec["segment_length_profile_km"]) + '"',
            ]
        )
    lines.extend(
        [
            'RW_RAMP_METER_IDS = "' + ",".join(str(r["id"]) for r in ramp_meters) + '"',
            'RW_RAMP_METER_SCS = "' + ",".join(str(int(r["sc_no"])) for r in ramp_meters) + '"',
            'RW_RAMP_METER_CONNECTORS = "' + ",".join(str(int(r["connector"])) for r in ramp_meters) + '"',
            'RW_RAMP_METER_MODEL_KEYS = "' + ",".join(str(r["model_ramp_key"]) for r in ramp_meters) + '"',
            'RW_SIGNAL_SCS = "' + ",".join(str(int(row["sc_no"])) for row in signals) + '"',
            'RW_SIGNAL_SG_FILTERS = "' + signal_sg_filter(rows) + '"',
            'RW_SIGNAL_HEAD_FILTERS = "' + signal_head_filter(rows) + '"',
            'RW_LOCAL_OBSERVABLE_LINKS = "' + ",".join(str(int(v)) for v in detector.get("observable_links", [])) + '"',
            f'RW_DETECTOR_MAPPING_PATH = "{rel(detector_path)}"',
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_wrapper(path: Path, tuning_path: Path, mapping_path: Path, generated_path: Path, network_path: Path) -> None:
    text = (ROOT / "scripts/run_real_world_single_watchdog.ps1").read_text(encoding="utf-8")
    text = text.replace(
        'evaluation\\configs\\real_world_modi_pstack_adapter_v0_20260719.json',
        rel(tuning_path),
    )
    text = text.replace(
        'evaluation\\real_world_modi_control\\control_mapping.json',
        rel(mapping_path),
    )
    text = text.replace(
        'network\\real_world_gaepo_modi\\modi_eval_rw_control.inpx',
        rel(network_path),
    )
    text = text.replace(
        'evaluation\\generated\\real_world_modi_control_config.vbs',
        rel(generated_path),
    )
    text = text.replace('[string]$Controller = "stackelberg"', '[string]$Controller = "pstack-flagship"')
    text = text.replace(
        rel(tuning_path),
        'evaluation\\configs\\real_world_modi_pstack_flagship_20260731.json',
    )
    path.write_text(text, encoding="utf-8")


def write_audit_csv(path: Path, rows: list[ExcelHead], controlled_uf_ids: set[int]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "urban_follower_id",
                "signal_id",
                "control_enabled",
                "sc_no",
                "sg_no",
                "movement",
                "phase",
                "axis",
                "signal_head_no",
                "link_no",
                "lane_no",
                "pos_m",
            ],
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "urban_follower_id": row.uf_id,
                    "signal_id": row.signal_id,
                    "control_enabled": row.uf_id in controlled_uf_ids,
                    "sc_no": row.sc_no,
                    "sg_no": row.sg_no,
                    "movement": row.movement,
                    "phase": row.phase,
                    "axis": axis_id(row.movement),
                    "signal_head_no": row.signal_head_no,
                    "link_no": row.link_no,
                    "lane_no": row.lane_no,
                    "pos_m": round(row.pos_m, 6),
                }
            )


def extra_head_report(rows: list[ExcelHead], network_path: Path) -> list[dict[str, Any]]:
    _, all_heads = parse_network(network_path)
    selected_heads_by_sc: dict[int, set[int]] = defaultdict(set)
    selected_sgs_by_sc: dict[int, set[int]] = defaultdict(set)
    for row in rows:
        selected_heads_by_sc[row.sc_no].add(row.signal_head_no)
        selected_sgs_by_sc[row.sc_no].add(row.sg_no)

    all_by_sc: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for head_no, spec in all_heads.items():
        sc_no = spec.get("sc_no")
        if sc_no is not None:
            all_by_sc[int(sc_no)].append({"head": head_no, **spec})

    out: list[dict[str, Any]] = []
    for sc_no, selected_heads in sorted(selected_heads_by_sc.items()):
        extras = [head for head in all_by_sc.get(sc_no, []) if int(head["head"]) not in selected_heads]
        if not extras:
            continue
        same_selected_sg = [head for head in extras if int(head.get("sg_no") or -1) in selected_sgs_by_sc[sc_no]]
        filtered_sg = [head for head in extras if int(head.get("sg_no") or -1) not in selected_sgs_by_sc[sc_no]]
        out.append(
            {
                "sc_no": sc_no,
                "filtered_extra_heads": len(filtered_sg),
                "same_sg_extra_heads": len(same_selected_sg),
                "same_sg_extra_head_sample": [head["head"] for head in same_selected_sg[:20]],
            }
        )
    return out


def write_report(
    path: Path,
    rows: list[ExcelHead],
    controlled_uf_ids: set[int],
    warnings: list[str],
    files: dict[str, Path],
    network_path: Path,
) -> None:
    grouped = group_heads(rows)
    controlled_rows = filtered_rows(rows, controlled_uf_ids)
    extras = extra_head_report(controlled_rows, network_path)
    lines = [
        "# Urban follower Excel adapter",
        "",
        f"Generated on 2026-07-31 from `Urban-Follower.xlsx` and `{rel(network_path)}`.",
        "",
        "## Controlled followers",
        "",
        "| UF | VISSIM SC | SG filter | signal heads | observed links |",
        "| ---: | ---: | --- | ---: | --- |",
    ]
    for uf_id, uf_rows in grouped.items():
        if uf_id not in controlled_uf_ids:
            continue
        lines.append(
            f"| {uf_id} | {uf_rows[0].sc_no} | "
            f"{','.join(str(v) for v in sorted({row.sg_no for row in uf_rows}))} | "
            f"{len(uf_rows)} | "
            f"{','.join(str(v) for v in sorted({row.link_no for row in uf_rows}))} |"
        )
    monitoring_ids = [uf_id for uf_id in grouped if uf_id not in controlled_uf_ids]
    if monitoring_ids:
        lines.extend(
            [
                "",
                "## Monitoring-only followers",
                "",
                "| UF | VISSIM SC | signal heads | observed links |",
                "| ---: | ---: | ---: | --- |",
            ]
        )
        for uf_id in monitoring_ids:
            uf_rows = grouped[uf_id]
            lines.append(
                f"| {uf_id} | {uf_rows[0].sc_no} | "
                f"{len(uf_rows)} | "
                f"{','.join(str(v) for v in sorted({row.link_no for row in uf_rows}))} |"
            )
    lines.extend(
        [
            "",
            "## Files",
            "",
        ]
    )
    for label, file_path in files.items():
        lines.append(f"- {label}: `{rel(file_path)}`")
    lines.extend(
        [
            "",
            "## Validation",
            "",
            "- All workbook-listed SC numbers and signal heads exist in `modi.inpx`.",
            "- The watchdog wrapper defaults to `network\\real_world_gaepo_modi\\modi.inpx`.",
            "- VBS runtime signal control is limited to the controlled UF subset and SG-filtered with `RW_SIGNAL_SG_FILTERS`.",
        ]
    )
    if extras:
        lines.extend(
            [
                "- Some selected SCs contain extra signal heads in `modi.inpx`: extra SGs outside the workbook are left outside COM control; extra heads sharing a selected SG cannot be separated by SG-level COM control.",
                "",
                "| SC | extra heads on filtered-out SGs | extra heads sharing selected SG | same-SG sample |",
                "| ---: | ---: | ---: | --- |",
            ]
        )
        for item in extras:
            lines.append(
                f"| {item['sc_no']} | {item['filtered_extra_heads']} | "
                f"{item['same_sg_extra_heads']} | "
                f"{','.join(str(v) for v in item['same_sg_extra_head_sample'])} |"
            )
    if warnings:
        lines.extend(["", "## Warnings", ""])
        for warning in warnings:
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Smoke command",
            "",
            "```powershell",
            ".\\scripts\\run_real_world_single_watchdog_urban_follower_excel.ps1 -Name uf_excel_smoke -SimPeriod 600 -ControlIntervalSec 60 -Controller diagnostic-signal-major -ForceStepwise",
            "```",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_outputs(excel_path: Path, network_path: Path, controlled_uf_ids: set[int]) -> dict[str, Path]:
    rows, warnings, _controllers = read_excel_heads(excel_path, network_path)
    all_uf_ids = set(group_heads(rows))
    missing_controlled = sorted(controlled_uf_ids - all_uf_ids)
    if missing_controlled:
        raise ValueError(f"controlled UF IDs are not present in the workbook: {missing_controlled}")
    controlled_rows = filtered_rows(rows, controlled_uf_ids)
    if not controlled_rows:
        raise ValueError("At least one controlled UF ID is required")
    mapping_path = OUT_DIR / f"control_mapping_{SLUG}_{DATE_TAG}.json"
    detector_path = OUT_DIR / f"detector_local_mapping_{SLUG}_{DATE_TAG}.json"
    player_path = OUT_DIR / f"player_config_{SLUG}_{DATE_TAG}.json"
    audit_csv_path = OUT_DIR / f"signal_head_mapping_{SLUG}_{DATE_TAG}.csv"

    base_mapping = read_json(BASE_MAPPING)
    base_detector = read_json(BASE_DETECTOR)
    network_override = build_network_override(controlled_rows)
    mapping = build_control_mapping(base_mapping, controlled_rows, rows, excel_path, network_path, detector_path)
    detector = build_detector_mapping(base_detector, rows, controlled_uf_ids)
    tuning = build_tuning_config(
        network_override,
        mapping_path,
        detector_path,
        controlled_rows,
        rows,
        excel_path,
        network_path,
    )
    grouped = group_heads(rows)
    controlled_grouped = group_heads(controlled_rows)
    player_config = {
        "schema_version": 1,
        "created_at": "2026-07-31",
        "mode": f"real_world_modi_{SLUG}",
        "source_network": str(network_path),
        "source_excel": str(excel_path),
        "controlled_urban_follower_ids": sorted(controlled_uf_ids),
        "monitoring_only_urban_follower_ids": [uf_id for uf_id in grouped if uf_id not in controlled_uf_ids],
        "controlled_signal_controllers": [group[0].sc_no for group in controlled_grouped.values()],
        "players": [
            {
                "id": "leader_network_manager",
                "kind": "leader",
                "state_scope": "whole network split into freeway links 2/26 and UF1-UF19 local signal observations",
                "controlled_followers": ["freeway_follower_real_world"]
                + [f"urban_follower_UF{uf_id}" for uf_id in controlled_grouped],
            },
            {
                "id": "freeway_follower_real_world",
                "kind": "freeway_follower",
                "physical_links": [26, 2],
                "vsl_segments": [seg["segment_id"] for seg in base_mapping.get("segments", [])],
                "ramp_meters": [meter["id"] for meter in base_mapping.get("ramp_meters", [])],
            },
        ]
        + [
            {
                "id": f"urban_follower_UF{uf_id}" if uf_id in controlled_uf_ids else f"urban_monitor_UF{uf_id}",
                "kind": "urban_follower" if uf_id in controlled_uf_ids else "urban_monitor",
                "signal_id": f"UF{uf_id}",
                "urban_follower_id": uf_id,
                "sc_no": group[0].sc_no,
                "control_enabled": uf_id in controlled_uf_ids,
                "monitoring_only": uf_id not in controlled_uf_ids,
                "local_observation_agent": f"U_UF{uf_id}",
                "signal_group_filter": sorted({row.sg_no for row in group}),
                "signal_heads": [row.signal_head_no for row in group],
            }
            for uf_id, group in grouped.items()
        ],
        "source_files": {
            "excel": str(excel_path),
            "network_inpx": str(network_path),
            "base_control_mapping": str(BASE_MAPPING),
            "base_detector_mapping": str(BASE_DETECTOR),
            "prediction_calibration": str(CALIBRATION),
        },
    }

    write_json(mapping_path, mapping)
    write_json(detector_path, detector)
    write_json(player_path, player_config)
    write_json(CONFIG_PATH, tuning)
    write_generated_vbs(GENERATED_VBS_PATH, mapping, detector, controlled_rows, detector_path)
    write_wrapper(WRAPPER_PATH, CONFIG_PATH, mapping_path, GENERATED_VBS_PATH, network_path)
    write_audit_csv(audit_csv_path, rows, controlled_uf_ids)
    write_report(
        REPORT_PATH,
        rows,
        controlled_uf_ids,
        warnings,
        {
            "control mapping": mapping_path,
            "detector mapping": detector_path,
            "player config": player_path,
            "audit csv": audit_csv_path,
            "P-Stack tuning": CONFIG_PATH,
            "generated VBS config": GENERATED_VBS_PATH,
            "watchdog wrapper": WRAPPER_PATH,
        },
        network_path,
    )
    return {
        "mapping": mapping_path,
        "detector": detector_path,
        "player_config": player_path,
        "audit_csv": audit_csv_path,
        "tuning": CONFIG_PATH,
        "generated_vbs": GENERATED_VBS_PATH,
        "wrapper": WRAPPER_PATH,
        "report": REPORT_PATH,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL))
    parser.add_argument("--network", default=str(DEFAULT_NETWORK))
    parser.add_argument(
        "--controlled-uf-ids",
        default=",".join(str(value) for value in DEFAULT_CONTROLLED_UF_IDS),
        help="Comma-separated Urban Follower IDs to actuate. Others remain monitoring-only.",
    )
    args = parser.parse_args()
    files = build_outputs(
        Path(args.excel),
        Path(args.network),
        parse_controlled_uf_ids(args.controlled_uf_ids),
    )
    for key, value in files.items():
        print(f"{key}={value}")


if __name__ == "__main__":
    main()
