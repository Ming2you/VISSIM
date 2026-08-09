from __future__ import annotations

import argparse
import csv
import json
import math
import re
import statistics
import xml.etree.ElementTree as ET
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATE_TAG = "20260731"

DEFAULT_EXCEL = ROOT / "Urban-Follower.xlsx"
DEFAULT_NETWORK = ROOT / "network/real_world_gaepo_modi/modi.inpx"
DEFAULT_CONTROLLED_UF_IDS = (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 13, 14, 15, 16)

OUT_DIR = ROOT / "evaluation/real_world_modi_control_urban_follower_excel_20260731"
CSV_PATH = OUT_DIR / f"midblock_signal_mapping_urban_follower_excel_{DATE_TAG}.csv"
JSON_PATH = OUT_DIR / f"midblock_signal_mapping_urban_follower_excel_{DATE_TAG}.json"
REPORT_PATH = ROOT / f"outputs/urban_follower_midblock_signal_mapping_{DATE_TAG}.md"


@dataclass(frozen=True)
class WorkbookHead:
    uf_id: int
    sc_no: int
    sg_no: int
    movement: str
    head_no: int


@dataclass(frozen=True)
class SignalHead:
    head_no: int
    sc_no: int
    sg_no: int
    link_no: int
    lane_no: int
    pos_m: float
    x: float | None
    y: float | None
    speed_kph: float | None


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT)).replace("/", "\\")
    except ValueError:
        return str(path)


def first_int(text: str | None) -> int | None:
    match = re.search(r"-?\d+", str(text or ""))
    return int(match.group(0)) if match else None


def parse_lane_ref(text: str | None) -> tuple[int, int]:
    nums = [int(value) for value in re.findall(r"-?\d+", str(text or ""))]
    if not nums:
        raise ValueError(f"Cannot parse lane reference: {text!r}")
    link = nums[0]
    lane = nums[1] if len(nums) > 1 else 1
    return link, lane


def parse_sg_ref(text: str | None) -> tuple[int, int] | None:
    nums = [int(value) for value in re.findall(r"-?\d+", str(text or ""))]
    if len(nums) < 2:
        return None
    return nums[0], nums[1]


def parse_movement(text: Any) -> tuple[int, str]:
    raw = str(text or "").strip()
    nums = [int(value) for value in re.findall(r"-?\d+", raw)]
    if not nums:
        raise ValueError(f"Cannot parse movement label from {raw!r}")
    tail = raw.split(":", 1)[1].strip() if ":" in raw else raw
    match = re.search(r"[A-Za-z]+", tail)
    movement = match.group(0).upper() if match else f"SG{nums[0]}"
    return nums[0], movement


def phase_id(label: str) -> str:
    value = str(label or "").upper()
    if value.startswith(("E", "W")) or "EB" in value or "WB" in value:
        return "p2"
    if value.startswith(("N", "S")) or "NB" in value or "SB" in value:
        return "p1"
    return ""


def main_sg_for(sg_no: int) -> int:
    return ((int(sg_no) - 1) % 8) + 1


def sg_block(sg_no: int) -> str:
    if 1 <= sg_no <= 8:
        return "main_1_8"
    if 9 <= sg_no <= 16:
        return "midblock_9_16"
    if 17 <= sg_no <= 24:
        return "midblock_17_24"
    return "outside_1_24_pattern"


def parse_controlled_uf_ids(text: str) -> set[int]:
    if not text.strip():
        return set(DEFAULT_CONTROLLED_UF_IDS)
    return {int(value) for value in re.findall(r"-?\d+", text)}


def read_workbook(path: Path) -> list[WorkbookHead]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    column = {name: idx for idx, name in enumerate(headers)}
    required = ["Urban Follower ID", "SC", "No.", "Signal Head"]
    missing = [name for name in required if name not in column]
    if missing:
        raise ValueError(f"{path} is missing required columns: {missing}")

    rows: list[WorkbookHead] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in values):
            continue
        if values[column["Urban Follower ID"]] is None or values[column["SC"]] is None:
            continue
        if values[column["No."]] is None or values[column["Signal Head"]] is None:
            continue
        sg_no, movement = parse_movement(values[column["No."]])
        rows.append(
            WorkbookHead(
                uf_id=int(float(values[column["Urban Follower ID"]])),
                sc_no=int(float(values[column["SC"]])),
                sg_no=sg_no,
                movement=movement,
                head_no=int(float(values[column["Signal Head"]])),
            )
        )
    return sorted(rows, key=lambda row: (row.uf_id, row.sg_no, row.head_no))


def link_coordinate(points: list[tuple[float, float]], pos_m: float) -> tuple[float, float] | None:
    if not points:
        return None
    if len(points) == 1:
        return points[0]
    remaining = max(0.0, float(pos_m))
    for start, end in zip(points, points[1:]):
        length = math.dist(start, end)
        if length <= 1.0e-9:
            continue
        if remaining <= length:
            ratio = remaining / length
            return (
                start[0] + ratio * (end[0] - start[0]),
                start[1] + ratio * (end[1] - start[1]),
            )
        remaining -= length
    return points[-1]


def parse_network(path: Path) -> tuple[dict[int, dict[str, Any]], dict[int, list[SignalHead]]]:
    root = ET.parse(path).getroot()
    controllers: dict[int, dict[str, Any]] = {}
    for sc in root.iter("signalController"):
        sc_no = first_int(sc.get("no"))
        if sc_no is None:
            continue
        controllers[sc_no] = {
            "name": sc.get("name", ""),
            "sg_names": {
                int(first_int(sg.get("no"))): sg.get("name", "")
                for sg in sc.findall("./sgs/signalGroup")
                if first_int(sg.get("no")) is not None
            },
        }

    links: dict[int, dict[str, Any]] = {}
    for link in root.iter("link"):
        link_no = first_int(link.get("no"))
        if link_no is None:
            continue
        points: list[tuple[float, float]] = []
        for point in link.findall("./geometry/linkPolyPts/linkPolyPoint"):
            try:
                points.append((float(point.get("x", "0")), float(point.get("y", "0"))))
            except ValueError:
                continue
        try:
            speed_kph = float(link.get("assumSpeedOncom", ""))
        except ValueError:
            speed_kph = None
        links[link_no] = {"points": points, "speed_kph": speed_kph}

    heads_by_sc: dict[int, list[SignalHead]] = defaultdict(list)
    for head in root.iter("signalHead"):
        head_no = first_int(head.get("no"))
        sg_ref = parse_sg_ref(head.get("sg"))
        if head_no is None or sg_ref is None:
            continue
        link_no, lane_no = parse_lane_ref(head.get("lane"))
        try:
            pos_m = float(head.get("pos", "0"))
        except ValueError:
            pos_m = 0.0
        link = links.get(link_no, {})
        xy = link_coordinate(link.get("points", []), pos_m)
        heads_by_sc[int(sg_ref[0])].append(
            SignalHead(
                head_no=head_no,
                sc_no=int(sg_ref[0]),
                sg_no=int(sg_ref[1]),
                link_no=link_no,
                lane_no=lane_no,
                pos_m=pos_m,
                x=xy[0] if xy else None,
                y=xy[1] if xy else None,
                speed_kph=link.get("speed_kph"),
            )
        )
    return controllers, dict(heads_by_sc)


def grouped_workbook(rows: list[WorkbookHead]) -> dict[int, list[WorkbookHead]]:
    out: dict[int, list[WorkbookHead]] = defaultdict(list)
    for row in rows:
        out[row.uf_id].append(row)
    return dict(sorted(out.items()))


def selected_heads_by_sg(rows: list[WorkbookHead]) -> dict[int, list[int]]:
    out: dict[int, list[int]] = defaultdict(list)
    for row in rows:
        out[row.sg_no].append(row.head_no)
    return dict(out)


def heads_by_sg(rows: list[SignalHead]) -> dict[int, list[SignalHead]]:
    out: dict[int, list[SignalHead]] = defaultdict(list)
    for row in rows:
        out[row.sg_no].append(row)
    return dict(sorted(out.items()))


def centroid(heads: list[SignalHead]) -> tuple[float, float] | None:
    coords = [(head.x, head.y) for head in heads if head.x is not None and head.y is not None]
    if not coords:
        return None
    return (sum(x for x, _ in coords) / len(coords), sum(y for _, y in coords) / len(coords))


def median_speed(heads: list[SignalHead]) -> float:
    values = [head.speed_kph for head in heads if head.speed_kph and head.speed_kph > 0]
    return float(statistics.median(values)) if values else 50.0


def round_to_5(value: float) -> int:
    return int(round(float(value) / 5.0) * 5)


def joined(values: list[Any] | set[Any]) -> str:
    return "|".join(str(value) for value in sorted(values))


def movement_for_sg(
    sc_no: int,
    sg_no: int,
    controller: dict[str, Any],
    workbook_by_sg: dict[int, list[WorkbookHead]],
) -> str:
    if sg_no in workbook_by_sg and workbook_by_sg[sg_no]:
        return workbook_by_sg[sg_no][0].movement
    return str(controller.get("sg_names", {}).get(sg_no, ""))


def build_mapping_rows(
    workbook_rows: list[WorkbookHead],
    controllers: dict[int, dict[str, Any]],
    heads_by_sc: dict[int, list[SignalHead]],
    controlled_uf_ids: set[int],
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for uf_id, uf_rows in grouped_workbook(workbook_rows).items():
        sc_no = uf_rows[0].sc_no
        controller = controllers.get(sc_no, {"name": "", "sg_names": {}})
        sc_heads = heads_by_sc.get(sc_no, [])
        by_sg = heads_by_sg(sc_heads)
        workbook_by_sg = defaultdict(list)
        for row in uf_rows:
            workbook_by_sg[row.sg_no].append(row)
        selected_head_nos = {row.head_no for row in uf_rows}
        selected_sg_nos = set(workbook_by_sg)
        parent_control_enabled = uf_id in controlled_uf_ids

        for sg_no, network_heads in by_sg.items():
            extra_heads = [head for head in network_heads if head.head_no not in selected_head_nos]
            if not extra_heads:
                continue

            if sg_no in selected_sg_nos:
                parent_sg = sg_no
                parent_heads = [head for head in network_heads if head.head_no in selected_head_nos]
                classification = "same_main_sg_extra_heads"
                recommendation = "already_tied_to_parent_sg_no_separate_offset"
            else:
                parent_sg = main_sg_for(sg_no)
                parent_heads = [
                    head
                    for head in by_sg.get(parent_sg, [])
                    if head.head_no in selected_heads_by_sg(uf_rows).get(parent_sg, [])
                ]
                if not parent_heads:
                    parent_heads = list(by_sg.get(parent_sg, []))
                classification = sg_block(sg_no)
                recommendation = "derive_offset_from_parent_uf" if parent_control_enabled else "monitoring_only_leave_existing"

            parent_label = movement_for_sg(sc_no, parent_sg, controller, workbook_by_sg)
            child_label = str(controller.get("sg_names", {}).get(sg_no, ""))
            phase = phase_id(parent_label) or phase_id(child_label)
            child_phase = phase_id(child_label)
            parent_xy = centroid(parent_heads)
            child_xy = centroid(extra_heads)
            distance_m = ""
            speed_kph = median_speed(extra_heads + parent_heads)
            delta_sec = ""
            delta_sec_round5 = ""
            if parent_xy and child_xy:
                distance = math.dist(parent_xy, child_xy)
                distance_m = round(distance, 3)
                delta = distance / max(1.0, speed_kph * 1000.0 / 3600.0)
                delta_sec = round(delta, 3)
                delta_sec_round5 = round_to_5(delta)

            parent_present = parent_sg in selected_sg_nos
            phase_match = not child_phase or not phase or child_phase == phase
            if classification == "same_main_sg_extra_heads":
                confidence = "high" if parent_heads else "medium"
                reason = "extra heads share a workbook-selected SG; SG-level COM cannot separate them"
            elif classification.startswith("midblock") and parent_present and phase_match and distance_m != "":
                confidence = "high"
                reason = "SG number folds to a workbook-selected 1-8 parent SG and geometry distance was computed"
            elif classification.startswith("midblock") and parent_present:
                confidence = "medium"
                reason = "SG number folds to a workbook-selected parent SG; geometry distance needs review"
            else:
                confidence = "review"
                reason = "outside selected parent pattern or parent SG is not in the workbook"

            out.append(
                {
                    "uf_id": uf_id,
                    "parent_signal_id": f"UF{uf_id}",
                    "parent_control_enabled": parent_control_enabled,
                    "sc_no": sc_no,
                    "sc_name": controller.get("name", ""),
                    "mid_sg_no": sg_no,
                    "mid_sg_name": child_label,
                    "classification": classification,
                    "parent_main_sg_no": parent_sg,
                    "parent_main_sg_name_or_movement": parent_label,
                    "parent_phase": phase,
                    "head_nos": joined({head.head_no for head in extra_heads}),
                    "link_nos": joined({head.link_no for head in extra_heads}),
                    "lane_nos": joined({head.lane_no for head in extra_heads}),
                    "head_count": len(extra_heads),
                    "parent_head_nos": joined({head.head_no for head in parent_heads}),
                    "distance_m_straight": distance_m,
                    "speed_kph_used": round(speed_kph, 3),
                    "delta_sec_raw": delta_sec,
                    "delta_sec_round5": delta_sec_round5,
                    "confidence": confidence,
                    "recommendation": recommendation,
                    "reason": reason,
                }
            )
    return sorted(out, key=lambda row: (int(row["uf_id"]), int(row["mid_sg_no"]), str(row["head_nos"])))


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fieldnames = [
        "uf_id",
        "parent_signal_id",
        "parent_control_enabled",
        "sc_no",
        "sc_name",
        "mid_sg_no",
        "mid_sg_name",
        "classification",
        "parent_main_sg_no",
        "parent_main_sg_name_or_movement",
        "parent_phase",
        "head_nos",
        "link_nos",
        "lane_nos",
        "head_count",
        "parent_head_nos",
        "distance_m_straight",
        "speed_kph_used",
        "delta_sec_raw",
        "delta_sec_round5",
        "confidence",
        "recommendation",
        "reason",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_json(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(rows, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")


def write_report(path: Path, rows: list[dict[str, Any]], controlled_uf_ids: set[int]) -> None:
    controlled = [row for row in rows if row["parent_control_enabled"]]
    monitoring = [row for row in rows if not row["parent_control_enabled"]]
    lines = [
        "# Urban follower midblock signal draft",
        "",
        f"Generated on 2026-07-31 from `Urban-Follower.xlsx` and `{rel(DEFAULT_NETWORK)}`.",
        "",
        "## Assumptions",
        "",
        "- SG 1-8 are treated as main-intersection groups.",
        "- SG 9-16 and SG 17-24 are treated as repeated midblock groups.",
        "- Parent main SG is computed as `((mid_sg_no - 1) mod 8) + 1`.",
        "- `delta_sec_round5` is a draft positive lag from straight-line stopline distance and the INPX link assumed speed; review before using it as a field offset.",
        "- Heads sharing a workbook-selected SG cannot be offset separately through SG-level COM control.",
        "",
        "## Summary",
        "",
        f"- Controlled UF IDs: {','.join(str(v) for v in sorted(controlled_uf_ids))}",
        f"- Draft rows: {len(rows)}",
        f"- Controlled-parent rows: {len(controlled)}",
        f"- Monitoring-only-parent rows: {len(monitoring)}",
        "",
        "## Draft Rows",
        "",
        "| UF | control | SC | mid SG | class | parent SG | phase | heads | links | delta s | confidence | recommendation |",
        "| ---: | :---: | ---: | ---: | --- | ---: | --- | --- | --- | ---: | --- | --- |",
    ]
    for row in rows:
        delta = row["delta_sec_round5"]
        head_nos = str(row["head_nos"]).replace("|", ",")
        link_nos = str(row["link_nos"]).replace("|", ",")
        lines.append(
            f"| {row['uf_id']} | {'Y' if row['parent_control_enabled'] else 'N'} | "
            f"{row['sc_no']} | {row['mid_sg_no']} | {row['classification']} | "
            f"{row['parent_main_sg_no']} | {row['parent_phase']} | {head_nos} | "
            f"{link_nos} | {delta} | {row['confidence']} | {row['recommendation']} |"
        )
    lines.extend(
        [
            "",
            "## Outputs",
            "",
            f"- CSV: `{rel(CSV_PATH)}`",
            f"- JSON: `{rel(JSON_PATH)}`",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL))
    parser.add_argument("--network", default=str(DEFAULT_NETWORK))
    parser.add_argument(
        "--controlled-uf-ids",
        default=",".join(str(value) for value in DEFAULT_CONTROLLED_UF_IDS),
    )
    parser.add_argument("--csv-out", default=str(CSV_PATH))
    parser.add_argument("--json-out", default=str(JSON_PATH))
    parser.add_argument("--report-out", default=str(REPORT_PATH))
    args = parser.parse_args()

    controlled_uf_ids = parse_controlled_uf_ids(args.controlled_uf_ids)
    workbook_rows = read_workbook(Path(args.excel))
    controllers, network_heads = parse_network(Path(args.network))
    rows = build_mapping_rows(workbook_rows, controllers, network_heads, controlled_uf_ids)
    write_csv(Path(args.csv_out), rows)
    write_json(Path(args.json_out), rows)
    write_report(Path(args.report_out), rows, controlled_uf_ids)
    print(f"rows={len(rows)}")
    print(f"csv={Path(args.csv_out)}")
    print(f"json={Path(args.json_out)}")
    print(f"report={Path(args.report_out)}")


if __name__ == "__main__":
    main()
